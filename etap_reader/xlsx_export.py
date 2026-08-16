"""
Excel (.xlsx) export helpers - generic table export and the aggregated
violations/alerts report. CSV export is handled client-side (no server
round-trip needed for plain text); this module exists because a real .xlsx
needs an actual library, not string-building.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CRITICAL_FILL = PatternFill("solid", fgColor="FFCDD2")


def _autosize(ws, columns, rows, sample=200):
    for i, col in enumerate(columns, 1):
        longest = len(str(col))
        for row in rows[:sample]:
            v = row.get(col)
            if v is not None:
                longest = max(longest, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(longest + 2, 45)


def _write_sheet(ws, columns, rows, highlight_col=None, highlight_values=("critical",)):
    ws.append(columns)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"

    highlight_idx = columns.index(highlight_col) if highlight_col in columns else None
    for row in rows:
        ws.append([row.get(c) for c in columns])
        if highlight_idx is not None and str(row.get(highlight_col, "")).lower() in highlight_values:
            for cell in ws[ws.max_row]:
                cell.fill = CRITICAL_FILL

    _autosize(ws, columns, rows)


def table_to_xlsx_bytes(columns, rows, sheet_name="Data"):
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Data")[:31]
    _write_sheet(ws, columns, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _unique_sheet_title(title, used):
    """Excel caps a sheet name at 31 characters and forbids []:*?/\\, and two
    sheets cannot share a name. ETAP table names are short enough that
    collisions are rare, but H1-H8 history variants of a long name do collide
    once truncated, and one collision fails the whole workbook."""
    clean = "".join("_" if c in "[]:*?/\\" else c for c in str(title))[:31] or "Sheet"
    if clean.lower() not in used:
        used.add(clean.lower())
        return clean
    for n in range(2, 1000):
        suffix = f"~{n}"
        candidate = clean[:31 - len(suffix)] + suffix
        if candidate.lower() not in used:
            used.add(candidate.lower())
            return candidate
    raise ValueError(f"could not find a free sheet name for {title!r}")


def all_tables_xlsx_bytes(sheet_source, index_columns, index_rows):
    """A workbook holding every table's contents, one sheet each.

    `sheet_source` yields (title, columns, rows) and is consumed lazily: a
    project model has 500 non-empty tables and a time-domain study has a
    million rows, so nothing here holds more than one table at a time.

    Written in openpyxl's write-only mode for the same reason - the normal
    mode keeps every cell as an object until save, which a million rows will
    not survive inside a container.

    The first sheet is the index, so a workbook of 500 tabs opens on something
    that says what is in it.
    """
    from openpyxl.cell import WriteOnlyCell  # noqa: PLC0415

    wb = Workbook(write_only=True)
    used = set()

    def header(ws, columns):
        cells = []
        for col in columns:
            c = WriteOnlyCell(ws, value=col)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            cells.append(c)
        ws.append(cells)

    index = wb.create_sheet(title=_unique_sheet_title("Index", used))
    index.freeze_panes = "A2"
    header(index, index_columns)
    for row in index_rows:
        index.append([row.get(c) for c in index_columns])

    for title, columns, rows in sheet_source:
        ws = wb.create_sheet(title=_unique_sheet_title(title, used))
        ws.freeze_panes = "A2"
        header(ws, columns)
        for row in rows:
            ws.append([row.get(c) for c in columns])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def multi_sheet_xlsx_bytes(sheets):
    """sheets: list of (title, columns, rows) tuples - one sheet each.
    Rows whose AlertType is "Critical" are highlighted."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, columns, rows in sheets:
        ws = wb.create_sheet(title=title[:31])
        _write_sheet(ws, columns, rows, highlight_col="AlertType")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
